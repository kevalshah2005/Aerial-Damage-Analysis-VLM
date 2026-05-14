#!/usr/bin/env python3
"""
xview2_dataset_report.py / dataset_analysis.py

Comprehensive xView2-style dataset report:
- Pair-based accounting: <disaster>_<index>_(pre|post)_disaster.json
- Buildings counted from PRE only (footprints)
- Damage counted from POST only (subtype)
- Severity distribution computed from POST only (ordinal mapping)
- Polygon geometry stats from PRE polygons (px^2; optional m^2 if gsd exists)
- Grouping by (location, disaster_type) where:
    disaster = "<location>-<disaster_type>"  (split by LAST hyphen)

Outputs:
- Terminal report (disable with --no_terminal)
- JSON report file (disable with --no_json)
- Excel report file (disable with --no_excel)

Usage:
  python3 dataset_analysis.py --input_dir /path/to/dataset_root
  python3 dataset_analysis.py --input_dir . --no_json
  python3 dataset_analysis.py --input_dir . --no_terminal
  python3 dataset_analysis.py --input_dir . --no_excel
"""

from __future__ import annotations

import argparse
import json
import math
import re
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Excel writer (openpyxl is installed in your environment)
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except Exception as e:
    Workbook = None  # type: ignore
    get_column_letter = None  # type: ignore
    _OPENPYXL_IMPORT_ERROR = e
else:
    _OPENPYXL_IMPORT_ERROR = None


# ----------------------------
# Severity mapping (POST only)
# ----------------------------
# 0 = none, 1 = minor, 2 = major, 3 = destroyed
SEVERITY_MAP = {
    "no-damage": 0,
    "minor-damage": 1,
    "major-damage": 2,
    "destroyed": 3,
}


# ----------------------------
# Filename parsing
# ----------------------------
# Accept both "disaster" and the common misspelling "diaster"
FNAME_RE = re.compile(
    r"^(?P<disaster>.+?)_(?P<idx>\d+)_(?P<phase>pre|post)_(?:disaster|diaster)\.json$",
    re.IGNORECASE
)

def parse_label_filename(name: str) -> Optional[Tuple[str, str, str]]:
    """Return (disaster, idx, phase) or None."""
    m = FNAME_RE.match(name)
    if not m:
        return None
    return (m.group("disaster"), m.group("idx"), m.group("phase").lower())

def split_location_type(disaster: str) -> Tuple[str, str]:
    """
    Split 'guatemala-volcano' -> ('guatemala', 'volcano') by LAST hyphen.
    If no hyphen, type = 'unknown'.
    """
    if "-" not in disaster:
        return disaster, "unknown"
    loc, typ = disaster.rsplit("-", 1)
    return loc, typ


# ----------------------------
# WKT polygon parsing (xy)
# ----------------------------
WKT_POLYGON_RE = re.compile(r"^\s*POLYGON\s*\(\(\s*(.*?)\s*\)\)\s*$", re.IGNORECASE)

def parse_wkt_polygon_xy(wkt: str) -> Optional[List[Tuple[float, float]]]:
    """
    Parse WKT:
      POLYGON ((x1 y1, x2 y2, ..., xN yN))
    Returns a closed list of points (first point repeated at end).
    """
    if not isinstance(wkt, str):
        return None
    m = WKT_POLYGON_RE.match(wkt.strip())
    if not m:
        return None

    pts: List[Tuple[float, float]] = []
    for pair in m.group(1).split(","):
        pair = pair.strip()
        if not pair:
            continue
        parts = pair.split()
        if len(parts) < 2:
            return None
        try:
            x = float(parts[0])
            y = float(parts[1])
        except ValueError:
            return None
        pts.append((x, y))

    if len(pts) < 3:
        return None

    if pts[0] != pts[-1]:
        pts.append(pts[0])
    return pts

def polygon_area_px2(points: List[Tuple[float, float]]) -> float:
    """Shoelace formula, absolute area in px^2. Assumes closed polygon."""
    if len(points) < 4:
        return 0.0
    s = 0.0
    for (x1, y1), (x2, y2) in zip(points[:-1], points[1:]):
        s += x1 * y2 - x2 * y1
    return abs(s) / 2.0

def polygon_perimeter_px(points: List[Tuple[float, float]]) -> float:
    """Perimeter in pixels. Assumes closed polygon."""
    if len(points) < 4:
        return 0.0
    p = 0.0
    for (x1, y1), (x2, y2) in zip(points[:-1], points[1:]):
        p += math.hypot(x2 - x1, y2 - y1)
    return p


# ----------------------------
# Data structures
# ----------------------------
@dataclass
class PairFiles:
    disaster: str
    idx: str
    location: str
    dtype: str
    pre: Optional[Path] = None
    post: Optional[Path] = None

@dataclass
class PairMetrics:
    disaster: str
    idx: str
    location: str
    dtype: str
    pre_label: str
    post_label: str
    buildings_pre: int
    buildings_post: int
    damage_counts_post: Dict[str, int]
    damaged_incl_unclassified: float
    damaged_excl_unclassified: float
    mean_severity_excl_unknown: Optional[float]
    # metadata snapshots (best-effort)
    width: Optional[int] = None
    height: Optional[int] = None
    sensor: Optional[str] = None
    gsd: Optional[float] = None
    off_nadir_angle: Optional[float] = None
    capture_date: Optional[str] = None  # raw string


# ----------------------------
# JSON helpers
# ----------------------------
def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))

def iter_buildings_xy(data: dict):
    """Yield dict entries from features.xy where properties.feature_type == 'building'."""
    feats = data.get("features", {})
    xy = feats.get("xy", [])
    if not isinstance(xy, list):
        return
    for e in xy:
        if not isinstance(e, dict):
            continue
        props = e.get("properties", {})
        if not isinstance(props, dict):
            continue
        if props.get("feature_type") != "building":
            continue
        yield e

def safe_get_metadata(d: dict) -> dict:
    md = d.get("metadata", {})
    return md if isinstance(md, dict) else {}

def parse_year_from_capture_date(s: Optional[str]) -> Optional[int]:
    if not s or not isinstance(s, str):
        return None
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.year
    except Exception:
        return None


# ----------------------------
# Statistics helpers
# ----------------------------
def quantile_sorted(vals_sorted: List[float], q: float) -> float:
    """
    Linear interpolation quantile on sorted list.
    q in [0,1].
    """
    n = len(vals_sorted)
    if n == 0:
        return float("nan")
    if n == 1:
        return vals_sorted[0]
    q = max(0.0, min(1.0, q))  # <-- FIXED (removed stray 'x')
    pos = (n - 1) * q
    lo = int(math.floor(pos))
    hi = int(math.ceil(pos))
    if lo == hi:
        return vals_sorted[lo]
    w = pos - lo
    return vals_sorted[lo] * (1 - w) + vals_sorted[hi] * w

def summarize_numeric(values: List[float], percentiles: List[int] = [0, 5, 10, 25, 50, 75, 90, 95, 99, 100]) -> Dict[str, Any]:
    """Returns dict with count, mean, stdev, min/max, and selected percentiles."""
    if not values:
        return {"count": 0}
    vs = sorted(values)
    out: Dict[str, Any] = {
        "count": len(vs),
        "mean": statistics.fmean(vs),
        "min": vs[0],
        "max": vs[-1],
    }
    out["stdev"] = statistics.pstdev(vs) if len(vs) >= 2 else 0.0
    out["percentiles"] = {f"p{p}": quantile_sorted(vs, p / 100.0) for p in percentiles}
    return out

def counts_with_percentages(cnt: Counter) -> Dict[str, Any]:
    total = sum(cnt.values())
    return {
        "total": total,
        "counts": dict(cnt),
        "percent": {k: (v / total if total else 0.0) for k, v in cnt.items()},
    }

def compute_damaged_rates(damage_cnt: Counter) -> Dict[str, float]:
    """
    Rates from POST counts:
      damaged_incl_unclassified = (total - no-damage) / total
      damaged_excl_unclassified = (total - no-damage - un-classified) / (total - un-classified)
    """
    total = sum(damage_cnt.values())
    no_damage = damage_cnt.get("no-damage", 0)
    uncls = damage_cnt.get("un-classified", 0)

    damaged_incl = (total - no_damage) / total if total else 0.0
    denom = (total - uncls)
    damaged_excl = (total - no_damage - uncls) / denom if denom > 0 else 0.0

    return {
        "damaged_incl_unclassified": damaged_incl,
        "damaged_excl_unclassified": damaged_excl,
    }

def summarize_severity(sev_scores: List[float], sev_counts: Counter) -> Dict[str, Any]:
    total = sum(sev_counts.values())
    out = {
        "mapping": {k: v for k, v in SEVERITY_MAP.items()},
        "counts": {str(k): int(v) for k, v in sev_counts.items()},
        "percent": {str(k): (v / total if total else 0.0) for k, v in sev_counts.items()},
        "scores": summarize_numeric(sev_scores) if sev_scores else {"count": 0},
        "rates": {
            "any_damage_rate_ge1": ((sev_counts.get(1, 0) + sev_counts.get(2, 0) + sev_counts.get(3, 0)) / total) if total else 0.0,
            "severe_damage_rate_ge2": ((sev_counts.get(2, 0) + sev_counts.get(3, 0)) / total) if total else 0.0,
        },
    }
    return out


# ----------------------------
# Reporting helpers
# ----------------------------
def fmt_int(n: int) -> str:
    return f"{n:,}"

def fmt_float(x: Optional[float], digits: int = 3) -> str:
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return "N/A"
    return f"{x:.{digits}f}"


# ----------------------------
# Excel helpers
# ----------------------------
def _excel_autosize(ws) -> None:
    if get_column_letter is None:
        return
    # basic autosize: based on max string length in column
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            s = str(v)
            max_len = max(max_len, len(s))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 60)

def _write_table(ws, headers: List[str], rows: List[List[Any]], freeze: bool = True) -> None:
    ws.append(headers)
    for r in rows:
        ws.append(r)
    if freeze:
        ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _excel_autosize(ws)

def save_excel_report(report: Dict[str, Any], out_path: Path) -> None:
    if Workbook is None:
        raise RuntimeError(f"openpyxl import failed: {_OPENPYXL_IMPORT_ERROR}")

    wb = Workbook()

    # --- Overview sheet
    ws0 = wb.active
    ws0.title = "overview"
    overview_rows = [
        ["generated_at", report["run_info"]["generated_at"]],
        ["input_dir", report["run_info"]["input_dir"]],
        ["labels_dir", report["run_info"]["labels_dir"]],
        ["labels_json_total", report["files"]["labels_json_total"]],
        ["labels_json_matched_pattern", report["files"]["labels_json_matched_pattern"]],
        ["labels_json_skipped", report["files"]["labels_json_skipped"]],
        ["bases_total", report["pairs"]["bases_total"]],
        ["complete_pairs", report["pairs"]["complete_pairs"]],
        ["missing_pre", report["pairs"]["missing_pre"]],
        ["missing_post", report["pairs"]["missing_post"]],
        ["total_buildings_pre", report["overall"]["buildings_pre_total"]],
        ["avg_buildings_per_pair", report["core"]["avg_buildings_per_pair"]],
    ]
    _write_table(ws0, ["key", "value"], overview_rows, freeze=False)

    # --- Overall damage sheet
    ws1 = wb.create_sheet("overall_damage")
    dmg = report["overall"]["damage_post_overall"]
    dmg_counts = dmg.get("counts", {})
    dmg_percent = dmg.get("percent", {})
    rows = []
    for subtype, c in sorted(dmg_counts.items(), key=lambda kv: kv[1], reverse=True):
        rows.append([subtype, c, float(dmg_percent.get(subtype, 0.0))])
    _write_table(ws1, ["subtype", "count", "percent"], rows)

    # --- Overall severity sheet
    ws2 = wb.create_sheet("overall_severity")
    sev = report["overall"].get("severity_post_overall", {})
    sev_counts = sev.get("counts", {})
    sev_percent = sev.get("percent", {})
    rows = []
    for k in ["0", "1", "2", "3"]:
        rows.append([k, sev_counts.get(k, 0), float(sev_percent.get(k, 0.0))])
    _write_table(ws2, ["severity_level", "count", "percent"], rows)

    # Add severity score summary
    ws2b = wb.create_sheet("severity_score_stats")
    sev_scores = sev.get("scores", {})
    rows2 = []
    for kk in ["count", "mean", "stdev", "min", "max"]:
        if kk in sev_scores:
            rows2.append([kk, sev_scores[kk]])
    perc = sev_scores.get("percentiles", {})
    for pk, pv in perc.items():
        rows2.append([pk, pv])
    _write_table(ws2b, ["metric", "value"], rows2, freeze=False)

    # --- Groups sheet (location, type)
    ws3 = wb.create_sheet("groups")
    group_rows = report["groups"]["by_location_type"]

    # union all damage subtypes across groups -> columns
    all_subtypes = set()
    for gr in group_rows:
        all_subtypes.update(gr["damage_post"]["counts"].keys())
    all_subtypes = set(sorted(all_subtypes))

    headers = [
        "location", "type", "pairs", "buildings_pre_total", "avg_buildings_per_pair",
        "damaged_rate_incl_unclassified", "damaged_rate_excl_unclassified",
        "mean_severity", "any_damage_rate_ge1", "severe_damage_rate_ge2",
        "poly_area_px2_mean", "poly_area_px2_median",
        "poly_area_m2_mean", "poly_area_m2_median",
    ]
    # add damage subtype columns
    for st in sorted(all_subtypes):
        headers.append(f"damage_{st}")

    rows = []
    for gr in group_rows:
        dmg_rates = gr["damage_post"]["rates"]
        sev_gr = gr.get("severity_post", {})
        sev_scores_gr = sev_gr.get("scores", {})
        sev_rates_gr = sev_gr.get("rates", {})
        px = gr.get("polygon_area_px2", {})
        px_p = px.get("percentiles", {})
        m2 = gr.get("polygon_area_m2", {})
        m2_p = m2.get("percentiles", {})

        row = [
            gr["location"],
            gr["type"],
            gr["pairs"],
            gr["buildings_pre_total"],
            gr["avg_buildings_per_pair"],
            float(dmg_rates.get("damaged_incl_unclassified", 0.0)),
            float(dmg_rates.get("damaged_excl_unclassified", 0.0)),
            sev_scores_gr.get("mean", None),
            float(sev_rates_gr.get("any_damage_rate_ge1", 0.0)),
            float(sev_rates_gr.get("severe_damage_rate_ge2", 0.0)),
            px.get("mean", None),
            px_p.get("p50", None),
            m2.get("mean", None),
            m2_p.get("p50", None),
        ]

        dmg_counts_gr = gr["damage_post"]["counts"]
        for st in sorted(all_subtypes):
            row.append(dmg_counts_gr.get(st, 0))
        rows.append(row)

    _write_table(ws3, headers, rows)

    # --- Top pairs by buildings
    ws4 = wb.create_sheet("top_pairs_buildings")
    top_pairs_b = report["top"]["pairs_by_buildings_pre"]
    headers = [
        "disaster", "idx", "location", "type", "buildings_pre", "buildings_post",
        "damaged_incl_unclassified", "damaged_excl_unclassified", "mean_severity_excl_unknown",
        "sensor", "gsd", "off_nadir_angle", "capture_date", "pre_label", "post_label"
    ]
    rows = []
    for pm in top_pairs_b:
        rows.append([
            pm["disaster"], pm["idx"], pm["location"], pm["dtype"],
            pm["buildings_pre"], pm["buildings_post"],
            pm["damaged_incl_unclassified"], pm["damaged_excl_unclassified"], pm.get("mean_severity_excl_unknown"),
            pm.get("sensor"), pm.get("gsd"), pm.get("off_nadir_angle"), pm.get("capture_date"),
            pm["pre_label"], pm["post_label"]
        ])
    _write_table(ws4, headers, rows)

    # --- Top pairs by damaged
    ws5 = wb.create_sheet("top_pairs_damaged")
    top_pairs_d = report["top"]["pairs_by_damaged_incl_unclassified"]
    rows = []
    for pm in top_pairs_d:
        rows.append([
            pm["disaster"], pm["idx"], pm["location"], pm["dtype"],
            pm["buildings_pre"], pm["buildings_post"],
            pm["damaged_incl_unclassified"], pm["damaged_excl_unclassified"], pm.get("mean_severity_excl_unknown"),
            pm.get("sensor"), pm.get("gsd"), pm.get("off_nadir_angle"), pm.get("capture_date"),
            pm["pre_label"], pm["post_label"]
        ])
    _write_table(ws5, headers, rows)

    # --- Metadata sensors
    ws6 = wb.create_sheet("metadata_sensors")
    sensors = report["metadata"].get("sensors", {})
    rows = [[k, v] for k, v in sorted(sensors.items(), key=lambda kv: kv[1], reverse=True)]
    _write_table(ws6, ["sensor", "count"], rows)

    # --- Metadata numeric summaries
    ws7 = wb.create_sheet("metadata_numeric")
    def _summ_row(name: str, block: Dict[str, Any]) -> List[Any]:
        if not block or block.get("count", 0) == 0:
            return [name, 0, None, None, None, None, None]
        p = block.get("percentiles", {})
        return [name, block.get("count"), block.get("mean"), p.get("p50"), p.get("p90"), block.get("min"), block.get("max")]

    rows = [
        _summ_row("gsd_m_per_px", report["metadata"].get("gsd", {})),
        _summ_row("off_nadir_angle", report["metadata"].get("off_nadir_angle", {})),
        _summ_row("width", report["metadata"].get("width", {})),
        _summ_row("height", report["metadata"].get("height", {})),
    ]
    _write_table(ws7, ["metric", "count", "mean", "p50", "p90", "min", "max"], rows)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ----------------------------
# Main analysis
# ----------------------------
def build_report(input_dir: Path, labels_subdir: str, top_n: int) -> Dict[str, Any]:
    labels_dir = input_dir / labels_subdir
    if not labels_dir.is_dir():
        raise SystemExit(f"labels directory not found: {labels_dir}")

    # 1) collect pair files
    pairs: Dict[Tuple[str, str], PairFiles] = {}
    skipped_files: List[str] = []

    all_label_files = sorted(labels_dir.glob("*.json"))
    for fp in all_label_files:
        parsed = parse_label_filename(fp.name)
        if not parsed:
            skipped_files.append(fp.name)
            continue
        disaster, idx, phase = parsed
        location, dtype = split_location_type(disaster)
        key = (disaster, idx)
        if key not in pairs:
            pairs[key] = PairFiles(disaster=disaster, idx=idx, location=location, dtype=dtype)
        if phase == "pre":
            pairs[key].pre = fp
        else:
            pairs[key].post = fp

    total_bases = len(pairs)
    complete_pairs = [p for p in pairs.values() if p.pre and p.post]
    missing_pre = sum(1 for p in pairs.values() if not p.pre)
    missing_post = sum(1 for p in pairs.values() if not p.post)

    # 2) aggregate metrics
    all_locations = set()
    all_types = set()

    buildings_pre_total = 0
    buildings_per_pair: List[float] = []

    # polygon stats from PRE
    poly_areas_px2: List[float] = []
    poly_areas_m2: List[float] = []
    poly_perimeters_px: List[float] = []
    poly_vertices: List[float] = []

    # group-wise aggregations (location, dtype)
    pairs_by_group = Counter()
    buildings_pre_by_group = Counter()
    buildings_per_pair_by_group = defaultdict(list)

    poly_area_px2_by_group = defaultdict(list)
    poly_area_m2_by_group = defaultdict(list)
    poly_vertices_by_group = defaultdict(list)

    # damage (POST) overall and by group
    damage_overall = Counter()
    damage_by_group = defaultdict(Counter)

    # severity (POST) overall and by group
    severity_counts_overall = Counter()
    severity_scores_overall: List[float] = []
    severity_counts_by_group = defaultdict(Counter)
    severity_scores_by_group = defaultdict(list)

    # metadata distributions (best-effort from PRE)
    sensor_counts = Counter()
    gsd_values: List[float] = []
    off_nadir_values: List[float] = []
    width_values: List[float] = []
    height_values: List[float] = []
    year_counts = Counter()

    # pair-level summaries (useful for "top N")
    pair_metrics: List[PairMetrics] = []

    # lightweight info / sanity (understanding)
    building_count_mismatch_pairs = 0

    for p in complete_pairs:
        all_locations.add(p.location)
        all_types.add(p.dtype)
        group_key = (p.location, p.dtype)
        pairs_by_group[group_key] += 1

        pre_data = load_json(p.pre)
        post_data = load_json(p.post)

        pre_md = safe_get_metadata(pre_data)
        width = pre_md.get("width") if isinstance(pre_md.get("width"), int) else None
        height = pre_md.get("height") if isinstance(pre_md.get("height"), int) else None
        sensor = pre_md.get("sensor") if isinstance(pre_md.get("sensor"), str) else None
        gsd = pre_md.get("gsd") if isinstance(pre_md.get("gsd"), (int, float)) else None
        off_nadir = pre_md.get("off_nadir_angle") if isinstance(pre_md.get("off_nadir_angle"), (int, float)) else None
        cap_date = pre_md.get("capture_date") if isinstance(pre_md.get("capture_date"), str) else None

        if sensor:
            sensor_counts[sensor] += 1
        if isinstance(gsd, (int, float)) and gsd > 0:
            gsd_values.append(float(gsd))
        if isinstance(off_nadir, (int, float)):
            off_nadir_values.append(float(off_nadir))
        if width is not None:
            width_values.append(float(width))
        if height is not None:
            height_values.append(float(height))
        yr = parse_year_from_capture_date(cap_date)
        if yr is not None:
            year_counts[yr] += 1

        # PRE: buildings + polygons
        pre_buildings = 0
        for e in iter_buildings_xy(pre_data):
            pre_buildings += 1
            wkt = e.get("wkt", "")
            pts = parse_wkt_polygon_xy(wkt) if isinstance(wkt, str) else None
            if pts:
                a_px2 = polygon_area_px2(pts)
                poly_areas_px2.append(a_px2)
                poly_area_px2_by_group[group_key].append(a_px2)

                poly_perimeters_px.append(polygon_perimeter_px(pts))

                vcount = max(0, len(pts) - 1)  # closed polygon repeats first point
                poly_vertices.append(float(vcount))
                poly_vertices_by_group[group_key].append(float(vcount))

                if isinstance(gsd, (int, float)) and gsd > 0:
                    a_m2 = a_px2 * (float(gsd) ** 2)
                    poly_areas_m2.append(a_m2)
                    poly_area_m2_by_group[group_key].append(a_m2)

        buildings_pre_total += pre_buildings
        buildings_per_pair.append(float(pre_buildings))
        buildings_pre_by_group[group_key] += pre_buildings
        buildings_per_pair_by_group[group_key].append(float(pre_buildings))

        # POST: damage distribution + severity (POST ONLY)
        post_buildings = 0
        damage_cnt = Counter()
        sev_scores_pair: List[float] = []

        for e in iter_buildings_xy(post_data):
            post_buildings += 1
            props = e.get("properties", {})
            subtype = "unknown"
            if isinstance(props, dict):
                subtype = props.get("subtype") or "unknown"

            damage_cnt[subtype] += 1

            # severity (only for known classes)
            if subtype in SEVERITY_MAP:
                sev = SEVERITY_MAP[subtype]
                severity_counts_overall[sev] += 1
                severity_scores_overall.append(float(sev))
                severity_counts_by_group[group_key][sev] += 1
                severity_scores_by_group[group_key].append(float(sev))
                sev_scores_pair.append(float(sev))

        # aggregate damage (POST ONLY)
        damage_overall.update(damage_cnt)
        damage_by_group[group_key].update(damage_cnt)

        if post_buildings != pre_buildings:
            building_count_mismatch_pairs += 1

        rates = compute_damaged_rates(damage_cnt)
        mean_sev_pair = statistics.fmean(sev_scores_pair) if sev_scores_pair else None

        pair_metrics.append(
            PairMetrics(
                disaster=p.disaster,
                idx=p.idx,
                location=p.location,
                dtype=p.dtype,
                pre_label=p.pre.name,
                post_label=p.post.name,
                buildings_pre=pre_buildings,
                buildings_post=post_buildings,
                damage_counts_post=dict(damage_cnt),
                damaged_incl_unclassified=rates["damaged_incl_unclassified"],
                damaged_excl_unclassified=rates["damaged_excl_unclassified"],
                mean_severity_excl_unknown=mean_sev_pair,
                width=width,
                height=height,
                sensor=sensor,
                gsd=float(gsd) if isinstance(gsd, (int, float)) else None,
                off_nadir_angle=float(off_nadir) if isinstance(off_nadir, (int, float)) else None,
                capture_date=cap_date,
            )
        )

    n_pairs = len(complete_pairs)

    # 3) compute summaries
    buildings_per_pair_summary = summarize_numeric(buildings_per_pair)
    poly_area_px2_summary = summarize_numeric(poly_areas_px2)
    poly_area_m2_summary = summarize_numeric(poly_areas_m2) if poly_areas_m2 else {"count": 0}
    poly_perimeter_px_summary = summarize_numeric(poly_perimeters_px)
    poly_vertices_summary = summarize_numeric(poly_vertices)

    damage_overall_block = counts_with_percentages(damage_overall)
    damage_overall_block["rates"] = compute_damaged_rates(damage_overall)

    severity_overall_block = summarize_severity(severity_scores_overall, severity_counts_overall)

    # group blocks
    group_rows: List[Dict[str, Any]] = []
    for (loc, typ), pair_ct in sorted(pairs_by_group.items(), key=lambda kv: (-kv[1], kv[0][0], kv[0][1])):
        gkey = (loc, typ)

        dmg = damage_by_group[gkey]
        dmg_block = counts_with_percentages(dmg)
        dmg_block["rates"] = compute_damaged_rates(dmg)

        sev_cnt = severity_counts_by_group[gkey]
        sev_scores = severity_scores_by_group[gkey]
        sev_block = summarize_severity(sev_scores, sev_cnt)

        row = {
            "location": loc,
            "type": typ,
            "pairs": pair_ct,
            "buildings_pre_total": buildings_pre_by_group[gkey],
            "avg_buildings_per_pair": (buildings_pre_by_group[gkey] / pair_ct) if pair_ct else 0.0,
            "buildings_per_pair": summarize_numeric(buildings_per_pair_by_group[gkey]),
            "polygon_area_px2": summarize_numeric(poly_area_px2_by_group[gkey]),
            "polygon_area_m2": summarize_numeric(poly_area_m2_by_group[gkey]) if poly_area_m2_by_group[gkey] else {"count": 0},
            "polygon_vertices": summarize_numeric(poly_vertices_by_group[gkey]),
            "damage_post": dmg_block,
            "severity_post": sev_block,
        }
        group_rows.append(row)

    # top-N useful slices
    pairs_by_buildings = sorted(pair_metrics, key=lambda pm: pm.buildings_pre, reverse=True)[:top_n]

    def damaged_count_incl(pm: PairMetrics) -> int:
        total = sum(pm.damage_counts_post.values())
        return total - pm.damage_counts_post.get("no-damage", 0)

    pairs_by_damaged = sorted(pair_metrics, key=damaged_count_incl, reverse=True)[:top_n]

    groups_by_pairs = sorted(group_rows, key=lambda r: r["pairs"], reverse=True)[:top_n]
    groups_by_buildings = sorted(group_rows, key=lambda r: r["buildings_pre_total"], reverse=True)[:top_n]
    groups_by_damaged_rate = sorted(
        group_rows,
        key=lambda r: r["damage_post"]["rates"]["damaged_excl_unclassified"],
        reverse=True
    )[:top_n]
    groups_by_mean_severity = sorted(
        group_rows,
        key=lambda r: (r["severity_post"]["scores"].get("mean", -1.0) if r["severity_post"]["scores"].get("count", 0) else -1.0),
        reverse=True
    )[:top_n]

    # metadata summaries
    metadata_block = {
        "sensors": dict(sensor_counts),
        "gsd": summarize_numeric(gsd_values) if gsd_values else {"count": 0},
        "off_nadir_angle": summarize_numeric(off_nadir_values) if off_nadir_values else {"count": 0},
        "width": summarize_numeric(width_values) if width_values else {"count": 0},
        "height": summarize_numeric(height_values) if height_values else {"count": 0},
        "capture_years": dict(year_counts),
    }

    report: Dict[str, Any] = {
        "run_info": {
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "input_dir": str(input_dir),
            "labels_dir": str(labels_dir),
        },
        "files": {
            "labels_json_total": len(all_label_files),
            "labels_json_matched_pattern": len(all_label_files) - len(skipped_files),
            "labels_json_skipped": len(skipped_files),
            "skipped_examples": skipped_files[:25],
        },
        "pairs": {
            "bases_total": total_bases,
            "complete_pairs": n_pairs,
            "missing_pre": missing_pre,
            "missing_post": missing_post,
            "pairs_with_pre_post_building_count_mismatch": building_count_mismatch_pairs,
        },
        "core": {
            "total_data_pieces_pairs": n_pairs,
            "total_buildings_pre": buildings_pre_total,
            "avg_buildings_per_pair": (buildings_pre_total / n_pairs) if n_pairs else 0.0,
            "damage_distribution_by_location_type": "See groups.by_location_type[*].damage_post (POST only)",
            "severity_distribution_overall": "See overall.severity_post_overall (POST only)",
            "average_polygon_size": "See overall.polygons.area_px2.mean (and area_m2.mean if available)",
            "all_locations": sorted(all_locations),
            "all_types": sorted(all_types),
        },
        "overall": {
            "buildings_pre_total": buildings_pre_total,
            "buildings_per_pair": buildings_per_pair_summary,
            "damage_post_overall": damage_overall_block,
            "severity_post_overall": severity_overall_block,
            "polygons": {
                "area_px2": poly_area_px2_summary,
                "area_m2": poly_area_m2_summary,
                "perimeter_px": poly_perimeter_px_summary,
                "vertices": poly_vertices_summary,
            },
        },
        "groups": {
            "locations": sorted(all_locations),
            "types": sorted(all_types),
            "by_location_type": group_rows,
        },
        "top": {
            "pairs_by_buildings_pre": [asdict(pm) for pm in pairs_by_buildings],
            "pairs_by_damaged_incl_unclassified": [asdict(pm) for pm in pairs_by_damaged],
            "groups_by_pairs": groups_by_pairs,
            "groups_by_buildings_pre": groups_by_buildings,
            "groups_by_damaged_rate_excl_unclassified": groups_by_damaged_rate,
            "groups_by_mean_severity": groups_by_mean_severity,
        },
        "metadata": metadata_block,
    }

    return report


def print_terminal_report(report: Dict[str, Any], top_n: int) -> None:
    files = report["files"]
    pairs = report["pairs"]
    overall = report["overall"]
    groups = report["groups"]
    meta = report["metadata"]

    print("\n==============================")
    print(" Dataset Report")
    print("==============================")
    print(f"Input:      {report['run_info']['input_dir']}")
    print(f"Labels dir: {report['run_info']['labels_dir']}")

    print("\n--- Files ---")
    print(f"Label JSON files:        {fmt_int(files['labels_json_total'])}")
    print(f"Matched pattern:         {fmt_int(files['labels_json_matched_pattern'])}")
    print(f"Skipped (non-matching):  {fmt_int(files['labels_json_skipped'])}")
    if files["labels_json_skipped"]:
        print(f"  examples: {', '.join(files['skipped_examples'][:8])}")

    print("\n--- Pairs (data pieces) ---")
    print(f"Bases total (disaster+idx): {fmt_int(pairs['bases_total'])}")
    print(f"Complete pairs (pre+post):  {fmt_int(pairs['complete_pairs'])}")
    print(f"Missing pre:                {fmt_int(pairs['missing_pre'])}")
    print(f"Missing post:               {fmt_int(pairs['missing_post'])}")
    print(f"PRE/POST building-count mismatches (info): {fmt_int(pairs['pairs_with_pre_post_building_count_mismatch'])}")

    print("\n--- Buildings (PRE footprints) ---")
    print(f"Total buildings (PRE): {fmt_int(overall['buildings_pre_total'])}")
    bpp = overall["buildings_per_pair"]
    if bpp.get("count", 0) > 0:
        p = bpp["percentiles"]
        print(f"Buildings per pair: mean={fmt_float(bpp['mean'], 3)}, stdev={fmt_float(bpp['stdev'], 3)}")
        print(f"  p50={fmt_float(p['p50'], 0)}, p90={fmt_float(p['p90'], 0)}, p99={fmt_float(p['p99'], 0)}  (min={fmt_float(bpp['min'],0)}, max={fmt_float(bpp['max'],0)})")

    print("\n--- Polygon geometry (PRE) ---")
    poly = overall["polygons"]
    area_px2 = poly["area_px2"]
    if area_px2.get("count", 0) > 0:
        ap = area_px2["percentiles"]
        print(
            f"Polygon area px^2: mean={fmt_float(area_px2['mean'], 2)}, "
            f"median={fmt_float(ap['p50'], 2)}, p90={fmt_float(ap['p90'], 2)}, p99={fmt_float(ap['p99'], 2)}"
        )
    area_m2 = poly.get("area_m2", {"count": 0})
    if area_m2.get("count", 0) > 0:
        am = area_m2["percentiles"]
        print(
            f"Polygon area m^2:  mean={fmt_float(area_m2['mean'], 2)}, "
            f"median={fmt_float(am['p50'], 2)}, p90={fmt_float(am['p90'], 2)}, p99={fmt_float(am['p99'], 2)}"
        )
    verts = poly["vertices"]
    if verts.get("count", 0) > 0:
        vp = verts["percentiles"]
        print(f"Vertices per polygon: mean={fmt_float(verts['mean'], 2)}, p50={fmt_float(vp['p50'],0)}, p90={fmt_float(vp['p90'],0)}, max={fmt_float(verts['max'],0)}")

    print("\n--- Damage distribution (POST) ---")
    dmg = overall["damage_post_overall"]
    rates = dmg["rates"]
    print(f"Total buildings (POST counted): {fmt_int(dmg['total'])}")
    print(f"Damaged rate (incl unclassified): {fmt_float(rates['damaged_incl_unclassified']*100, 2)}%")
    print(f"Damaged rate (excl unclassified): {fmt_float(rates['damaged_excl_unclassified']*100, 2)}%")
    cnt = Counter(dmg["counts"])
    for k, v in cnt.most_common(6):
        print(f"  {k:15s} {fmt_int(v)}")

    print("\n--- Severity distribution (POST, ordinal) ---")
    sev = overall.get("severity_post_overall", {})
    sev_scores = sev.get("scores", {})
    if sev_scores.get("count", 0) > 0:
        sp = sev_scores["percentiles"]
        print("Mapping: 0=no-damage, 1=minor, 2=major, 3=destroyed")
        print(f"Mean severity: {fmt_float(sev_scores['mean'], 3)} | median: {fmt_float(sp['p50'], 0)} | p90: {fmt_float(sp['p90'], 0)}")
        print(f"Any-damage rate (>=1): {fmt_float(sev['rates']['any_damage_rate_ge1']*100, 2)}%")
        print(f"Severe-damage rate (>=2): {fmt_float(sev['rates']['severe_damage_rate_ge2']*100, 2)}%")
        for level in ["0", "1", "2", "3"]:
            print(f"  severity {level}: {fmt_int(int(sev['counts'].get(level, 0)))}")
    else:
        print("No severity scores available.")

    print("\n--- Groups ---")
    print(f"Locations: {len(groups['locations'])} | Types: {len(groups['types'])}")
    print(f"Top {top_n} (location, type) by pairs:")
    for row in report["top"]["groups_by_pairs"][:top_n]:
        loc = row["location"]
        typ = row["type"]
        pr = row["pairs"]
        btot = row["buildings_pre_total"]
        dr = row["damage_post"]["rates"]["damaged_excl_unclassified"]
        ms = row["severity_post"]["scores"].get("mean", None)
        print(
            f"  {loc:18s} | {typ:10s} | pairs={fmt_int(pr):>6s} | "
            f"bld_pre={fmt_int(btot):>8s} | damaged%={fmt_float(dr*100,2)} | mean_sev={fmt_float(ms,3)}"
        )

    print("\n--- Metadata (high-level) ---")
    sensors = meta.get("sensors", {})
    if sensors:
        top_s = Counter(sensors).most_common(5)
        print("Top sensors:")
        for s, c in top_s:
            print(f"  {s:18s} {fmt_int(c)}")
    gsd = meta.get("gsd", {"count": 0})
    if gsd.get("count", 0) > 0:
        print(f"GSD (m/px): mean={fmt_float(gsd['mean'], 4)}, p50={fmt_float(gsd['percentiles']['p50'],4)}, p90={fmt_float(gsd['percentiles']['p90'],4)}")
    off = meta.get("off_nadir_angle", {"count": 0})
    if off.get("count", 0) > 0:
        print(f"Off-nadir angle: mean={fmt_float(off['mean'],2)}, p50={fmt_float(off['percentiles']['p50'],2)}, p90={fmt_float(off['percentiles']['p90'],2)}")

    print("\nDone.\n")


def main():
    ap = argparse.ArgumentParser(description="Comprehensive xView2 dataset report (terminal + JSON + Excel).")
    ap.add_argument("--input_dir", type=str, required=True, help="Dataset root folder (contains labels/)")
    ap.add_argument("--labels_subdir", type=str, default="labels", help="Labels subfolder name (default: labels)")

    ap.add_argument("--json_path", type=str, default="", help="Where to save JSON report (default: <input_dir>/dataset_report.json)")
    ap.add_argument("--excel_path", type=str, default="", help="Where to save Excel report (default: <input_dir>/dataset_report.xlsx)")

    ap.add_argument("--top_n", type=int, default=10, help="Top-N items shown + stored in report (default: 10)")
    ap.add_argument("--no_json", action="store_true", help="Disable saving JSON report")
    ap.add_argument("--no_excel", action="store_true", help="Disable saving Excel report")
    ap.add_argument("--no_terminal", action="store_true", help="Disable terminal output")
    args = ap.parse_args()

    input_dir = Path(args.input_dir).expanduser().resolve()
    if not input_dir.is_dir():
        raise SystemExit(f"input_dir is not a directory: {input_dir}")

    report = build_report(input_dir=input_dir, labels_subdir=args.labels_subdir, top_n=args.top_n)

    if not args.no_terminal:
        print_terminal_report(report, top_n=args.top_n)

    if not args.no_json:
        out_json = Path(args.json_path).expanduser() if args.json_path else (input_dir / "dataset_report.json")
        out_json.parent.mkdir(parents=True, exist_ok=True)
        out_json.write_text(json.dumps(report, indent=2), encoding="utf-8")
        if not args.no_terminal:
            print(f"JSON report saved to: {out_json}")

    if not args.no_excel:
        out_xlsx = Path(args.excel_path).expanduser() if args.excel_path else (input_dir / "dataset_report.xlsx")
        save_excel_report(report, out_xlsx)
        if not args.no_terminal:
            print(f"Excel report saved to: {out_xlsx}")


if __name__ == "__main__":
    main()
